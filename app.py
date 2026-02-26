
import streamlit as st
import pandas as pd
import numpy as np
import re, json, math, io, uuid
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

APP_VERSION = "row-based-v16-global-price-row-lock"

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
MAPPING_DIR = DATA_DIR / "mappings"
MAPPING_DIR.mkdir(parents=True, exist_ok=True)
HISTORY_DIR = DATA_DIR / "history"
HISTORY_DIR.mkdir(parents=True, exist_ok=True)
DEFAULT_STANDARD_STOCKS_CSV = DATA_DIR / "standard_stocks.csv"

# Session bundle holds per-sheet calculated results to apply later
if "bundle" not in st.session_state:
    st.session_state.bundle = {}  # {sheet_name: {"lines": df}}
if "_force_restore" not in st.session_state:
    st.session_state._force_restore = False
if "_skip_restore_once" not in st.session_state:
    st.session_state._skip_restore_once = False

# ---------- helpers ----------
def clean_text(s) -> str:
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("Ø", "dia ").replace("⌀", "dia ")
    return s

def load_mapping(customer: str) -> dict:
    fp = MAPPING_DIR / f"{clean_text(customer).replace(' ', '_')}_stock_map.json"
    if fp.exists():
        return json.loads(fp.read_text(encoding="utf-8"))
    return {"customer": customer, "mappings": {}}

def save_mapping(customer: str, mapping: dict) -> None:
    fp = MAPPING_DIR / f"{clean_text(customer).replace(' ', '_')}_stock_map.json"
    fp.write_text(json.dumps(mapping, indent=2), encoding="utf-8")


def serialize_bundle(bundle: dict) -> dict:
    """
    Convert session bundle (dfs) into JSON-serializable structure.
    NOTE: This stores calculated line_totals and settings, not the original workbook.
    """
    out = {}
    for sh, data in (bundle or {}).items():
        settings = data.get("settings", {}) if isinstance(data, dict) else {}
        df = data.get("lines") if isinstance(data, dict) else None
        if df is None:
            continue
        try:
            lines_records = df.to_dict(orient="records")
        except Exception:
            lines_records = []
        out[sh] = {"settings": settings, "lines": lines_records}
    return out

def deserialize_bundle(obj: dict) -> dict:
    """
    Convert JSON structure back into session bundle (dfs).
    """
    bundle = {}
    if not isinstance(obj, dict):
        return bundle
    for sh, data in obj.items():
        if not isinstance(data, dict):
            continue
        settings = data.get("settings", {})
        lines_records = data.get("lines", [])
        try:
            df = pd.DataFrame(lines_records)
        except Exception:
            df = pd.DataFrame()
        bundle[sh] = {"settings": settings, "lines": df}
    return bundle


def _history_path(quote_id: str) -> Path:
    return HISTORY_DIR / f"{quote_id}.json"

def save_history_record(record: dict) -> None:
    quote_id = record.get("quote_id") or str(uuid.uuid4())
    record["quote_id"] = quote_id
    _history_path(quote_id).write_text(json.dumps(record, indent=2, default=str), encoding="utf-8")

def list_history_records() -> list[dict]:
    records = []
    for fp in sorted(HISTORY_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            records.append(json.loads(fp.read_text(encoding="utf-8")))
        except Exception:
            continue
    return records

def update_history_record(quote_id: str, updates: dict) -> None:
    fp = _history_path(quote_id)
    if not fp.exists():
        return
    try:
        rec = json.loads(fp.read_text(encoding="utf-8"))
    except Exception:
        return
    rec.update(updates)
    fp.write_text(json.dumps(rec, indent=2, default=str), encoding="utf-8")

def parse_size_text(text: str):
    """
    Supports:
      - "1200 x 900"
      - "H1700 x W800mm" (letter adjacent)
      - "1125W X 508 Hmm (2pcs)" (suffix)
      - "585mm x W662.5mm"
      - "DIA 600", "diameter 600", "450 round"
    """
    t = clean_text(text)
    if not t:
        return {"shape":"unknown","width_mm":np.nan,"height_mm":np.nan,"diameter_mm":np.nan}

    # remove bracket notes like (2pcs)
    t = re.sub(r"\([^)]*\)", " ", t)
    t = t.replace("mm", " ")
    t = re.sub(r"\s+", " ", t).strip()

    # circle
    m = re.search(r"(dia(?:meter)?)\s*[:=]?\s*(\d+(?:\.\d+)?)", t)
    if m:
        d = float(m.group(2))
        return {"shape":"circle","width_mm":np.nan,"height_mm":np.nan,"diameter_mm":d}
    m = re.search(r"(\d+(?:\.\d+)?)\s*round", t)
    if m:
        d = float(m.group(1))
        return {"shape":"circle","width_mm":np.nan,"height_mm":np.nan,"diameter_mm":d}

    # labelled width/height (any order), including adjacent patterns: w800, 800w, h1700, 1700h
    w = None
    h = None

    # width patterns
    mw = re.search(r"(?:\bwidth\b\s*[:=]?\s*(\d+(?:\.\d+)?))", t)
    if mw:
        w = float(mw.group(1))
    if w is None:
        mw = re.search(r"w\s*(\d+(?:\.\d+)?)", t)  # w800 or w 800
        if mw:
            w = float(mw.group(1))
    if w is None:
        mw = re.search(r"(\d+(?:\.\d+)?)\s*w", t)  # 800w
        if mw:
            w = float(mw.group(1))

    # height patterns
    mh = re.search(r"(?:\bheight\b\s*[:=]?\s*(\d+(?:\.\d+)?))", t)
    if mh:
        h = float(mh.group(1))
    if h is None:
        mh = re.search(r"h\s*(\d+(?:\.\d+)?)", t)  # h1700 or h 1700
        if mh:
            h = float(mh.group(1))
    if h is None:
        mh = re.search(r"(\d+(?:\.\d+)?)\s*h", t)  # 1700h
        if mh:
            h = float(mh.group(1))

    if w is not None and h is not None:
        return {"shape":"rectangle","width_mm":w,"height_mm":h,"diameter_mm":np.nan}

    # generic rectangle: "1200 x 900"
    m = re.search(r"(\d+(?:\.\d+)?)\s*(x|\*|by)\s*(\d+(?:\.\d+)?)", t)
    if m:
        a = float(m.group(1))
        b = float(m.group(3))
        # if we found one of w/h, use it to decide which is missing
        if w is not None and h is None:
            return {"shape":"rectangle","width_mm":w,"height_mm":b,"diameter_mm":np.nan}
        if h is not None and w is None:
            return {"shape":"rectangle","width_mm":a,"height_mm":h,"diameter_mm":np.nan}
        return {"shape":"rectangle","width_mm":a,"height_mm":b,"diameter_mm":np.nan}

    # fallback: first two numbers
    nums = re.findall(r"\d+(?:\.\d+)?", t)
    if len(nums) >= 2:
        return {"shape":"rectangle","width_mm":float(nums[0]),"height_mm":float(nums[1]),"diameter_mm":np.nan}

    return {"shape":"unknown","width_mm":np.nan,"height_mm":np.nan,"diameter_mm":np.nan}

def sqm_calc(shape: str, width_mm=None, height_mm=None, diameter_mm=None):
    if shape == "rectangle" and pd.notna(width_mm) and pd.notna(height_mm):
        return (float(width_mm)/1000.0) * (float(height_mm)/1000.0)
    if shape == "circle" and pd.notna(diameter_mm):
        r = (float(diameter_mm)/1000.0)/2.0
        return math.pi * (r**2)
    return np.nan

def sides_normalize(val: str, default="SS"):
    t = clean_text(val)
    if t in ("ds","2s","double","2pp","two","double sided","double-sided"):
        return "DS"
    if t in ("ss","1s","single","1pp","one","single sided","single-sided"):
        return "SS"
    return default

def eval_qty_value(qty_cell_val, ws):
    """
    If qty is a number -> return it.
    If qty is a simple Excel SUM formula like "=SUM(AB13:AB56)" -> compute it from sheet cells.
    Otherwise -> try numeric conversion.
    """
    if qty_cell_val is None:
        return np.nan

    # direct number
    if isinstance(qty_cell_val, (int, float)) and not (isinstance(qty_cell_val, float) and np.isnan(qty_cell_val)):
        return float(qty_cell_val)

    s = str(qty_cell_val).strip()
    # try numeric string
    n = pd.to_numeric(s, errors="coerce")
    if pd.notna(n):
        return float(n)

    # SUM(range)
    m = re.match(r"^=\s*sum\s*\(\s*([A-Z]+\d+)\s*:\s*([A-Z]+\d+)\s*\)\s*$", s, flags=re.IGNORECASE)
    if m:
        a1, b1 = m.group(1).upper(), m.group(2).upper()
        try:
            cells = ws[a1:b1]
            total = 0.0
            for row in cells:
                for c in row:
                    v = c.value
                    vv = pd.to_numeric(v, errors="coerce")
                    if pd.notna(vv):
                        total += float(vv)
            return total
        except Exception:
            return np.nan

    return np.nan

def export_quote_pdf(df_summary: pd.DataFrame, df_lines: pd.DataFrame, title="Quote") -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm

    bio = io.BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    y = h - 18*mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(18*mm, y, title)
    y -= 7*mm
    c.setFont("Helvetica", 9)
    c.drawString(18*mm, y, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {APP_VERSION}")
    y -= 8*mm

    c.setFont("Helvetica-Bold", 11)
    c.drawString(18*mm, y, "Summary")
    y -= 6*mm
    c.setFont("Helvetica", 10)
    for _, row in df_summary.iterrows():
        c.drawString(20*mm, y, f"{row['Label']}: {row['Value']}")
        y -= 5*mm

    y -= 4*mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(18*mm, y, "Line Items (Material)")
    y -= 6*mm

    headers = ["Col", "Qty", "Sides", "Size", "Stock", "Total SQM", "Rate", "DS%", "Line Total"]
    col_x = [18, 30, 42, 60, 98, 150, 170, 182, 198]  # mm
    c.setFont("Helvetica-Bold", 7)
    for hx, head in zip(col_x, headers):
        c.drawString(hx*mm, y, head)
    y -= 4*mm
    c.setFont("Helvetica", 7)

    def money(x):
        try: return f"{float(x):,.2f}"
        except: return str(x)

    for _, r in df_lines.head(100).iterrows():
        if y < 18*mm:
            c.showPage()
            y = h - 18*mm
            c.setFont("Helvetica-Bold", 7)
            for hx, head in zip(col_x, headers):
                c.drawString(hx*mm, y, head)
            y -= 4*mm
            c.setFont("Helvetica", 7)

        vals = [
            str(r.get("col_letter","")),
            str(int(r.get("qty",0))),
            str(r.get("sides","")),
            str(r.get("size_text",""))[:16],
            str(r.get("stock_std",""))[:22],
            money(r.get("total_sqm",0)),
            money(r.get("sqm_rate",0)),
            str(int(round((float(r.get("ds_factor",1.0))-1.0)*100))) if pd.notna(r.get("ds_factor")) else "0",
            money(r.get("line_total",0)),
        ]
        for hx, v in zip(col_x, vals):
            c.drawString(hx*mm, y, v)
        y -= 4*mm

    c.showPage()
    c.save()
    return bio.getvalue()

# ---------- UI ----------
st.set_page_config(page_title="Quoting App (Row-based)", layout="wide")
st.title("Quoting App — Row-based Input (Pick ROWS + Column Range)")
st.caption("Example: Size row 5, Material row 6, Sides row 10, Qty row 57. Each COLUMN = one line item.")

with st.sidebar:
    st.header("Upload")
    page = st.radio("Page", ["Quote Builder", "History"], index=0)

    customer = st.text_input("Customer", value="AU Holiday")
    uploaded = st.file_uploader("Customer Excel", type=["xlsx"])

    st.divider()
    st.header("Bundle")
    if st.button("Clear bundle"):
        st.session_state.bundle = {}
        st.success("Bundle cleared.")

    st.subheader("Export / Import bundle settings")
    # Export bundle JSON
    try:
        _bundle_payload = {
            "app_version": APP_VERSION,
            "customer": customer,
            "exported_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "price_row": int(st.session_state.get("price_row", 0)) if "price_row" in st.session_state else None,
            "bundle": serialize_bundle(st.session_state.bundle),
        }
        _bundle_json = json.dumps(_bundle_payload, indent=2)
        st.download_button(
            "Download Bundle JSON",
            data=_bundle_json.encode("utf-8"),
            file_name=f"bundle_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
            mime="application/json",
        )
    except Exception as e:
        st.caption("Bundle JSON export not available yet.")

    # Import bundle JSON
    bundle_file = st.file_uploader("Import Bundle JSON", type=["json"], key="bundle_json_upload")
    if bundle_file is not None:
        try:
            payload = json.loads(bundle_file.getvalue().decode("utf-8"))
            new_bundle = deserialize_bundle(payload.get("bundle", {}))
            st.session_state.bundle = new_bundle
            st.success(f"Imported bundle with {len(new_bundle)} sheet(s).")
        except Exception as e:
            st.error("Could not import this JSON file. Make sure it is a bundle exported from this app.")

    st.divider()
    st.header("Standard stock rates")
    st.caption("CSV columns required: stock_name_std, sqm_rate")
    std_upload = st.file_uploader("Optional: upload standard_stocks.csv", type=["csv"])

if uploaded is None:
    st.info("Upload an Excel file to start.")
    st.stop()

_uploaded_bytes = uploaded.getvalue()
_uploaded_name = getattr(uploaded, "name", "uploaded.xlsx")

# Standard stocks
if std_upload is not None:
    df_std = pd.read_csv(std_upload)
else:
    if DEFAULT_STANDARD_STOCKS_CSV.exists():
        df_std = pd.read_csv(DEFAULT_STANDARD_STOCKS_CSV)
    else:
        df_std = pd.DataFrame(columns=["stock_name_std","sqm_rate"])

df_std["stock_name_std"] = df_std["stock_name_std"].astype(str)
df_std["sqm_rate"] = pd.to_numeric(df_std["sqm_rate"], errors="coerce")
std_options = df_std["stock_name_std"].dropna().tolist()
std_rate_map = dict(zip(df_std["stock_name_std"], df_std["sqm_rate"]))


# ---------- HISTORY PAGE ----------
if page == "History":
    st.title("Quote History — Win/Lose")
    records = list_history_records()
    if not records:
        st.info("No history yet. Go to 'Quote Builder' and click 'Save Quote to History'.")
        st.stop()

    df = pd.DataFrame(records)

    f1, f2, f3, f4 = st.columns([1.2, 1.2, 1.2, 2.4])
    status_vals = sorted(df.get("status", pd.Series(["Pending"])).fillna("Pending").unique().tolist())
    status_filter = f1.multiselect("Status", options=status_vals, default=status_vals)

    cust_vals = sorted(df.get("customer", pd.Series(dtype=str)).fillna("").unique().tolist())
    customer_filter = f2.multiselect("Customer", options=cust_vals, default=cust_vals[: min(10, len(cust_vals))])

    date_contains = f3.text_input("Date contains", value="")
    search = f4.text_input("Search (file / notes / sheet)", value="")

    def _contains(s, needle):
        try:
            return needle.lower() in str(s).lower()
        except Exception:
            return False

    fdf = df.copy()
    if "status" in fdf.columns:
        fdf = fdf[fdf["status"].fillna("Pending").isin(status_filter)]
    if "customer" in fdf.columns and customer_filter:
        fdf = fdf[fdf["customer"].fillna("").isin(customer_filter)]
    if date_contains.strip():
        fdf = fdf[fdf.get("created_at", "").apply(lambda x: _contains(x, date_contains.strip()))]
    if search.strip():
        needle = search.strip()
        fdf = fdf[
            fdf.get("file_name", "").apply(lambda x: _contains(x, needle))
            | fdf.get("notes", "").apply(lambda x: _contains(x, needle))
            | fdf.get("applied_sheets", "").astype(str).apply(lambda x: _contains(x, needle))
        ]

    display_cols = ["quote_id","created_at","customer","file_name","status","loss_reason","sell_price","subtotal_material","total_sqm","ds_loading_pct","price_row","applied_sheets","notes"]
    display_cols = [c for c in display_cols if c in fdf.columns]
    edit_df = fdf[display_cols].copy()

    st.caption("Edit status/reason/notes then click 'Save changes'.")
    edited = st.data_editor(
        edit_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "status": st.column_config.SelectboxColumn("Status", options=["Pending","Won","Lost"]),
            "loss_reason": st.column_config.SelectboxColumn("Loss reason", options=["","Price","Timing","Spec/Capability","Service","Other"]),
        }
    )

    if st.button("Save changes", type="primary"):
        for _, row in edited.iterrows():
            qid = row.get("quote_id")
            if not qid:
                continue
            updates = {
                "status": row.get("status","Pending"),
                "loss_reason": row.get("loss_reason",""),
                "notes": row.get("notes",""),
                "sell_price": row.get("sell_price", None),
            }
            update_history_record(str(qid), updates)
        st.success("History updated.")
        st.rerun()

    st.download_button(
        "Download History CSV",
        data=fdf.to_csv(index=False).encode("utf-8"),
        file_name=f"quote_history_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv",
    )

    st.subheader("Analytics")
    c1, c2, c3 = st.columns(3)
    total = len(df)
    won = int((df.get("status","") == "Won").sum()) if "status" in df.columns else 0
    lost = int((df.get("status","") == "Lost").sum()) if "status" in df.columns else 0
    c1.metric("Total quotes", total)
    c2.metric("Won", won)
    c3.metric("Win rate", f"{(won/total*100):.1f}%" if total else "0.0%")

    if "loss_reason" in df.columns:
        st.caption("Loss reasons (count)")
        lr = df[df.get("status","") == "Lost"]["loss_reason"].fillna("").replace("", "Unspecified").value_counts()
        if len(lr):
            st.bar_chart(lr)

    st.stop()

# ---------- QUOTE BUILDER PAGE ----------
# Sheet names
uploaded.seek(0)
wb_ro = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
sheet_names = wb_ro.sheetnames
wb_ro.close()

top1, top2 = st.columns([2,1])
sheet_name = top1.selectbox("Sheet", sheet_names, index=0)

# Restore saved settings for this sheet BEFORE widgets are created (Streamlit requirement)
_saved = st.session_state.bundle.get(sheet_name, {}).get("settings")
_do_restore = bool(_saved) and (st.session_state.get("auto_restore", True) or st.session_state.get("_force_restore", False))
if st.session_state.get("_skip_restore_once", False):
    _do_restore = False
    st.session_state._skip_restore_once = False

if _do_restore:
    # Set widget default states BEFORE the widgets are instantiated below
    st.session_state["size_row"] = int(_saved.get("size_row", 5))
    st.session_state["mat_row"] = int(_saved.get("mat_row", 6))
    st.session_state["sides_row"] = int(_saved.get("sides_row", 10))
    st.session_state["qty_row"] = int(_saved.get("qty_row", 57))
    st.session_state["start_col_letter"] = str(_saved.get("start_col_letter", "A"))
    st.session_state["end_col_letter"] = str(_saved.get("end_col_letter", "Z"))
    st.session_state["units"] = str(_saved.get("units", "mm"))
    st.session_state["ds_loading_pct_pct"] = float(_saved.get("ds_loading_pct", 0.20)) * 100.0
    st.session_state["global_price_row"] = int(_saved.get("price_row", int(_saved.get("qty_row",57))+1))
    st.session_state["skip_zero_qty"] = bool(_saved.get("skip_zero_qty", True))
    st.session_state["_force_restore"] = False
units = top2.selectbox("Units", ["mm","cm","m"], index=0, key="units")
auto_save_on_open = st.checkbox("Auto-save this sheet to bundle when opened", value=False)

u = {"mm":1.0,"cm":10.0,"m":1000.0}[units]

# Preview (best-effort)
with st.expander("Preview (all rows) — best effort", expanded=False):
    try:
        uploaded.seek(0)
        df_preview = pd.read_excel(uploaded, sheet_name=sheet_name, header=None)
        st.dataframe(df_preview, use_container_width=True)
    except Exception as e:
        st.warning("Preview could not be rendered (sheet may be highly formatted). Row-based extraction will still work.")

st.subheader("Pick ROW numbers (1-indexed)")

c1, c2, c3, c4, c5 = st.columns([1.1,1.1,1.1,1.1,1.6])
size_row = c1.number_input("Size row", 1, 5000, 5, 1, key="size_row")
mat_row  = c2.number_input("Material row", 1, 5000, 6, 1, key="mat_row")
sides_row= c3.number_input("Sides row", 1, 5000, 10, 1, key="sides_row")
qty_row  = c4.number_input("Qty row", 1, 5000, 57, 1, key="qty_row")
default_sides = c5.radio("Default sides (if blank)", ["SS","DS"], index=0, horizontal=True, key="default_sides")

ds_loading_pct = st.number_input("DS loading (%)", min_value=0.0, max_value=100.0, value=20.0, step=1.0, key="ds_loading_pct_pct") / 100.0

st.subheader("Pick COLUMN range")
r1, r2, r3 = st.columns([1.2,1.2,1.6])
start_col_letter = r1.text_input("Start column letter", value="A", key="start_col_letter")
end_col_letter   = r2.text_input("End column letter", value="Z", key="end_col_letter")
skip_zero_qty    = r3.checkbox("Skip columns with Qty <= 0", value=True, key="skip_zero_qty")

st.subheader("Export output location")
st.caption("By default, price is written **next to the Qty row** (Qty row + 1).")
if "global_price_row" not in st.session_state:
    # initialize once (do NOT change when switching sheets)
    st.session_state["global_price_row"] = int(st.session_state.get("qty_row", int(qty_row))) + 1
global_price_row = st.number_input("Write price into row (GLOBAL, 1-indexed)", 1, 5000, int(st.session_state["global_price_row"]), 1, key="global_price_row")
# keep old variable name for minimal changes below
price_row = int(global_price_row)


apply_mode = st.radio("Apply prices to", ["Bundle sheets (saved)", "Current sheet only", "ALL sheets (same settings)"], index=0, horizontal=True, key="apply_mode")
all_sheets_selected = []
if apply_mode == "ALL sheets (same settings)":
    all_sheets_selected = st.multiselect("Select sheets to update", options=sheet_names, default=sheet_names, key="all_sheets_selected")
write_zero_when_missing_rate = st.checkbox("Write $0.00 when rate/mapping missing (otherwise leave blank)", value=False, key="write_zero_when_missing_rate")

# ---------- Extract row-based items ----------
uploaded.seek(0)
wb = openpyxl.load_workbook(uploaded, read_only=False, data_only=False)
ws = wb[sheet_name]

def col_idx(letter: str) -> int:
    letter = (letter or "").strip().upper()
    return column_index_from_string(letter)

c_start = col_idx(start_col_letter)
c_end = col_idx(end_col_letter)
if c_start > c_end:
    c_start, c_end = c_end, c_start

items = []
for c in range(c_start, c_end+1):
    size_val = ws.cell(row=int(size_row), column=c).value
    mat_val  = ws.cell(row=int(mat_row), column=c).value
    sides_val= ws.cell(row=int(sides_row), column=c).value
    qty_val  = ws.cell(row=int(qty_row), column=c).value

    qty = pd.to_numeric(eval_qty_value(qty_val, ws), errors="coerce")
    if skip_zero_qty and (pd.isna(qty) or float(qty) <= 0):
        continue

    size_text = "" if size_val is None else str(size_val)
    geo = parse_size_text(size_text)
    shape = geo["shape"] if geo["shape"] != "unknown" else "rectangle"

    width_mm = geo["width_mm"] * u if pd.notna(geo["width_mm"]) else np.nan
    height_mm = geo["height_mm"] * u if pd.notna(geo["height_mm"]) else np.nan
    diameter_mm = geo["diameter_mm"] * u if pd.notna(geo["diameter_mm"]) else np.nan

    sides = sides_normalize(sides_val, default=default_sides)

    items.append({
        "origin_col": c,
        "col_letter": get_column_letter(c),
        "qty": float(qty) if pd.notna(qty) else 0.0,
        "sides": sides,
        "size_text": size_text,
        "shape": shape,
        "width_mm": width_mm,
        "height_mm": height_mm,
        "diameter_mm": diameter_mm,
        "stock_customer": "" if mat_val is None else str(mat_val),
    })

lines = pd.DataFrame(items)
if len(lines) == 0:
    st.warning("No line items found in the selected column range (check Qty row and column range).")
    st.stop()

lines["sqm_each"] = lines.apply(lambda r: sqm_calc(r["shape"], r["width_mm"], r["height_mm"], r["diameter_mm"]), axis=1)
lines["total_sqm"] = pd.to_numeric(lines["sqm_each"], errors="coerce") * pd.to_numeric(lines["qty"], errors="coerce")

if pd.to_numeric(lines["sqm_each"], errors="coerce").notna().sum() == 0:
    st.error("Size parsing failed: SQM is blank for all columns. Check the Size row values and Units. Supported: W x H, H/W labels (H1700 x W800), DIA 600, 450 round.")

# ---------- Mapping ----------
st.subheader("Stock Mapping (Customer name → Standard stock with sqm rate)")
mapping = load_mapping(customer)
unique_cs = sorted({s for s in lines["stock_customer"].dropna().astype(str).tolist() if clean_text(s)})

map_df = pd.DataFrame([{
    "customer_stock": cs,
    "standard_stock": mapping["mappings"].get(clean_text(cs), "")
} for cs in unique_cs])

edited = st.data_editor(
    map_df,
    use_container_width=True,
    hide_index=True,
    column_config={
        "standard_stock": st.column_config.SelectboxColumn("Standard stock", options=[""] + std_options)
    }
)

if st.button("Save mappings", type="primary"):
    new_map = load_mapping(customer)
    for _, r in edited.iterrows():
        cs_key = clean_text(r["customer_stock"])
        std = str(r["standard_stock"] or "").strip()
        if cs_key and std:
            new_map["mappings"][cs_key] = std
    save_mapping(customer, new_map)
    st.success("Mappings saved.")

mapping = load_mapping(customer)
lines["stock_std"] = lines["stock_customer"].apply(lambda x: mapping["mappings"].get(clean_text(x), ""))
lines["sqm_rate"] = lines["stock_std"].map(std_rate_map)

# DS loading factor
lines["ds_factor"] = np.where(lines["sides"].astype(str) == "DS", 1.0 + ds_loading_pct, 1.0)

lines["line_total"] = (
    pd.to_numeric(lines["total_sqm"], errors="coerce")
    * pd.to_numeric(lines["sqm_rate"], errors="coerce")
    * pd.to_numeric(lines["ds_factor"], errors="coerce")
)

# ---------- Review ----------
st.subheader("Quote Review")
review = lines[["col_letter","qty","sides","ds_factor","shape","size_text","stock_customer","stock_std","sqm_each","total_sqm","sqm_rate","line_total"]].copy()
st.dataframe(review, use_container_width=True)

# ----- Bundle controls -----
st.subheader("Bundle (multi-sheet)")
bcol1, bcol2, bcol3 = st.columns([1.6, 1.6, 3.0])
with bcol1:
    if st.button("Save this sheet to bundle", type="primary"):
        st.session_state.bundle[sheet_name] = {"lines": lines.copy(), "settings": {
    "size_row": int(size_row),
    "mat_row": int(mat_row),
    "sides_row": int(sides_row),
    "qty_row": int(qty_row),
    "start_col_letter": str(start_col_letter).strip().upper(),
    "end_col_letter": str(end_col_letter).strip().upper(),
    "units": units,
    "ds_loading_pct": float(ds_loading_pct),
    "price_row": int(st.session_state.get("global_price_row", price_row)),
    "skip_zero_qty": bool(skip_zero_qty),
}}
        st.success(f"Saved '{sheet_name}' to bundle.")
with bcol2:
    if st.button("Remove this sheet from bundle"):
        st.session_state._skip_restore_once = True
        if sheet_name in st.session_state.bundle:
            del st.session_state.bundle[sheet_name]
            st.success(f"Removed '{sheet_name}' from bundle.")
        else:
            st.info("This sheet is not in the bundle.")
        st.rerun()
with bcol3:
    st.write("Sheets in bundle:", ", ".join(st.session_state.bundle.keys()) if st.session_state.bundle else "(none)")

st.caption("Bundle settings summary")
if st.session_state.bundle:
    rows = []
    for sh, data in st.session_state.bundle.items():
        s = (data or {}).get("settings", {}) if isinstance(data, dict) else {}
        rows.append({
            "sheet": sh,
            "size_row": s.get("size_row"),
            "mat_row": s.get("mat_row"),
            "sides_row": s.get("sides_row"),
            "qty_row": s.get("qty_row"),
            "cols": f"{s.get('start_col_letter','') or ''}:{s.get('end_col_letter','') or ''}",
            "units": s.get("units"),
            "ds_loading_%": round(float(s.get("ds_loading_pct", 0.0))*100.0, 2) if s.get("ds_loading_pct") is not None else None,
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

if auto_save_on_open and sheet_name not in st.session_state.bundle:
    st.session_state.bundle[sheet_name] = {"lines": lines.copy(), "settings": {
    "size_row": int(size_row),
    "mat_row": int(mat_row),
    "sides_row": int(sides_row),
    "qty_row": int(qty_row),
    "start_col_letter": str(start_col_letter).strip().upper(),
    "end_col_letter": str(end_col_letter).strip().upper(),
    "units": units,
    "ds_loading_pct": float(ds_loading_pct),
    "price_row": int(st.session_state.get("global_price_row", price_row)),
    "skip_zero_qty": bool(skip_zero_qty),
}}
    st.info(f"Auto-saved '{sheet_name}' to bundle.")

total_sqm = float(pd.to_numeric(review["total_sqm"], errors="coerce").fillna(0).sum())
subtotal = float(pd.to_numeric(review["line_total"], errors="coerce").fillna(0).sum())

summary = pd.DataFrame([
    {"Label":"Customer", "Value": customer},
    {"Label":"Sheet", "Value": sheet_name},
    {"Label":"Total SQM", "Value": f"{total_sqm:,.3f}"},
    {"Label":"Subtotal (Material)", "Value": f"{subtotal:,.2f}"},
    {"Label":"DS loading %", "Value": f"{ds_loading_pct*100:.0f}%"},
])

st.markdown("**Totals**")
st.table(summary)

def export_preserving_excel_all_sheets() -> bytes:
    """
    Preserve original formatting. Always write prices into the SAME global price_row.

    Modes:
    - Bundle sheets (saved): RE-CALCULATE each saved sheet using its saved settings (so each sheet can have different rows/cols/units).
    - Current sheet only: use current computed lines.
    - ALL sheets (same settings): recompute every selected sheet using current settings.
    """
    wb = openpyxl.load_workbook(io.BytesIO(_uploaded_bytes), read_only=False, data_only=False)

    apply_mode_local = st.session_state.get("apply_mode", "Bundle sheets (saved)")
    selected_local = st.session_state.get("all_sheets_selected", [])
    write_zero_missing = bool(st.session_state.get("write_zero_when_missing_rate", False))
    diag_rows = []

    # Load mappings once
    mapping = load_mapping(customer)

    def normalize_settings(s: dict) -> dict:
        s = s or {}
        # accept synonyms from older bundles/templates
        out = {}
        out["size_row"] = s.get("size_row", s.get("size_r", s.get("row_size", size_row)))
        out["mat_row"] = s.get("mat_row", s.get("material_row", s.get("stock_row", mat_row)))
        out["sides_row"] = s.get("sides_row", s.get("side_row", sides_row))
        out["qty_row"] = s.get("qty_row", s.get("quantity_row", qty_row))
        out["start_col_letter"] = s.get("start_col_letter", s.get("start_col", s.get("col_start", start_col_letter)))
        out["end_col_letter"] = s.get("end_col_letter", s.get("end_col", s.get("col_end", end_col_letter)))
        out["units"] = s.get("units", units)
        out["ds_loading_pct"] = s.get("ds_loading_pct", ds_loading_pct)
        out["skip_zero_qty"] = s.get("skip_zero_qty", skip_zero_qty)
        return out

    def compute_lines_for_sheet(sh: str, settings: dict) -> pd.DataFrame:
        if sh not in wb.sheetnames:
            return pd.DataFrame()
        ws_local = wb[sh]

        # settings (normalized)
        ns = normalize_settings(settings)
        size_r = int(ns.get("size_row", size_row))
        mat_r = int(ns.get("mat_row", mat_row))
        sides_r = int(ns.get("sides_row", sides_row))
        qty_r = int(ns.get("qty_row", qty_row))
        start_col = str(ns.get("start_col_letter", start_col_letter)).strip().upper()
        end_col = str(ns.get("end_col_letter", end_col_letter)).strip().upper()
        units_local = str(ns.get("units", units))
        ds_local = float(ns.get("ds_loading_pct", ds_loading_pct))
        skip_zero_local = bool(ns.get("skip_zero_qty", skip_zero_qty))

        u_local = {"mm":1.0,"cm":10.0,"m":1000.0}.get(units_local, 1.0)

        def col_idx(letter: str) -> int:
            letter = (letter or "").strip().upper()
            return column_index_from_string(letter)

        c_start = col_idx(start_col)
        c_end = col_idx(end_col)
        if c_start > c_end:
            c_start, c_end = c_end, c_start

        items = []
        for c in range(c_start, c_end + 1):
            size_val = ws_local.cell(row=size_r, column=c).value
            mat_val  = ws_local.cell(row=mat_r, column=c).value
            sides_val= ws_local.cell(row=sides_r, column=c).value
            qty_val  = ws_local.cell(row=qty_r, column=c).value

            qty = pd.to_numeric(eval_qty_value(qty_val, ws_local), errors="coerce")
            if skip_zero_local and (pd.isna(qty) or float(qty) <= 0):
                continue

            size_text = "" if size_val is None else str(size_val)
            geo = parse_size_text(size_text)
            shape = geo["shape"] if geo["shape"] != "unknown" else "rectangle"

            width_mm = geo["width_mm"] * u_local if pd.notna(geo["width_mm"]) else np.nan
            height_mm = geo["height_mm"] * u_local if pd.notna(geo["height_mm"]) else np.nan
            diameter_mm = geo["diameter_mm"] * u_local if pd.notna(geo["diameter_mm"]) else np.nan

            sides_norm = sides_normalize(sides_val, default=default_sides)

            items.append({
                "origin_col": c,
                "col_letter": get_column_letter(c),
                "qty": float(qty) if pd.notna(qty) else 0.0,
                "sides": sides_norm,
                "size_text": size_text,
                "shape": shape,
                "width_mm": width_mm,
                "height_mm": height_mm,
                "diameter_mm": diameter_mm,
                "stock_customer": "" if mat_val is None else str(mat_val),
            })

        df = pd.DataFrame(items)
        if len(df) == 0:
            diag_rows.append({
                "sheet": sh,
                "items": 0,
                "missing_rate_or_mapping": 0,
                "cols": f"{start_col}:{end_col}",
                "size_row": size_r,
                "mat_row": mat_r,
                "sides_row": sides_r,
                "qty_row": qty_r,
                "units": units_local,
                "ds_loading_pct": ds_local,
                "total_sqm": 0.0,
                "subtotal": 0.0,
            })
            return df

        df["sqm_each"] = df.apply(lambda r: sqm_calc(r["shape"], r["width_mm"], r["height_mm"], r["diameter_mm"]), axis=1)
        df["total_sqm"] = pd.to_numeric(df["sqm_each"], errors="coerce") * pd.to_numeric(df["qty"], errors="coerce")

        df["stock_std"] = df["stock_customer"].apply(lambda x: mapping["mappings"].get(clean_text(x), ""))
        df["sqm_rate"] = df["stock_std"].map(std_rate_map)
        missing_map = (df["stock_std"].fillna("") == "") | (df["sqm_rate"].isna())
        missing_count = int(missing_map.sum())
        if write_zero_missing and missing_count > 0:
            df.loc[missing_map, "sqm_rate"] = 0.0
        df["ds_factor"] = np.where(df["sides"].astype(str) == "DS", 1.0 + ds_local, 1.0)
        df["line_total"] = (
            pd.to_numeric(df["total_sqm"], errors="coerce")
            * pd.to_numeric(df["sqm_rate"], errors="coerce")
            * pd.to_numeric(df["ds_factor"], errors="coerce")
        )
        diag_rows.append({
            "sheet": sh,
            "items": int(len(df)),
            "missing_rate_or_mapping": int(((df["stock_std"].fillna("") == "") | (df["sqm_rate"].isna())).sum()) if "sqm_rate" in df.columns else 0,
            "total_sqm": float(pd.to_numeric(df.get("total_sqm", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()),
            "subtotal": float(pd.to_numeric(df.get("line_total", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()),
            "size_row": size_r,
            "mat_row": mat_r,
            "sides_row": sides_r,
            "qty_row": qty_r,
            "cols": f"{start_col}:{end_col}",
            "units": units_local,
            "ds_loading_pct": ds_local,
        })
        return df

    # Decide which sheets to apply
    to_apply = {}  # sh -> df_lines
    applied_settings = {}  # sh -> settings

    if apply_mode_local == "Bundle sheets (saved)" and st.session_state.bundle:
        for sh, data in st.session_state.bundle.items():
            settings = data.get("settings", {}) if isinstance(data, dict) else {}
            df_lines = compute_lines_for_sheet(sh, settings)
            to_apply[sh] = df_lines
            applied_settings[sh] = settings
    elif apply_mode_local == "Current sheet only":
        to_apply = {sheet_name: lines.copy()}
        applied_settings[sheet_name] = {
            "size_row": int(size_row), "mat_row": int(mat_row), "sides_row": int(sides_row), "qty_row": int(qty_row),
            "start_col_letter": str(start_col_letter).strip().upper(), "end_col_letter": str(end_col_letter).strip().upper(),
            "units": units, "ds_loading_pct": float(ds_loading_pct), "skip_zero_qty": bool(skip_zero_qty),
        }
    elif apply_mode_local == "ALL sheets (same settings)":
        settings = {
            "size_row": int(size_row), "mat_row": int(mat_row), "sides_row": int(sides_row), "qty_row": int(qty_row),
            "start_col_letter": str(start_col_letter).strip().upper(), "end_col_letter": str(end_col_letter).strip().upper(),
            "units": units, "ds_loading_pct": float(ds_loading_pct), "skip_zero_qty": bool(skip_zero_qty),
        }
        selected = selected_local if selected_local else wb.sheetnames
        for sh in selected:
            to_apply[sh] = compute_lines_for_sheet(sh, settings)
            applied_settings[sh] = settings
    else:
        to_apply = {sheet_name: lines.copy()}

    # Write values
    applied_sheets = []
    for sh, df_lines in to_apply.items():
        if sh not in wb.sheetnames or df_lines is None or len(df_lines) == 0:
            continue
        ws = wb[sh]
        for _, r in df_lines.iterrows():
            c = int(r["origin_col"])
            val = r.get("line_total")
            if pd.isna(val):
                if write_zero_missing:
                    val = 0.0
                else:
                    continue
            cell = ws.cell(row=int(price_row), column=c)
            cell.value = float(val)
            cell.number_format = "$#,##0.00"
        applied_sheets.append(sh)

    # Refresh summary sheets
    for name in ["Quote Summary", "Line Items"]:
        if name in wb.sheetnames:
            del wb[name]

    all_rows = []
    for sh, df_lines in to_apply.items():
        if df_lines is None or len(df_lines) == 0:
            continue
        df2 = df_lines.copy()
        df2["sheet"] = sh
        all_rows.append(df2)
    all_df = pd.concat(all_rows, ignore_index=True) if all_rows else pd.DataFrame()

    total_sqm_all = float(pd.to_numeric(all_df.get("total_sqm", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if len(all_df) else 0.0
    subtotal_all = float(pd.to_numeric(all_df.get("line_total", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if len(all_df) else 0.0

    ws_sum = wb.create_sheet("Quote Summary")
    sum_rows = [
        ("Customer", customer),
        ("Applied sheets", ", ".join(applied_sheets) if applied_sheets else "(none)"),
        ("Apply mode", apply_mode_local),
        ("Price row", int(price_row)),
        ("DS loading %", f"{ds_loading_pct*100:.0f}%"),
        ("Total SQM", f"{total_sqm_all:,.3f}"),
        ("Subtotal (Material)", f"{subtotal_all:,.2f}"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("App version", APP_VERSION),
    ]
    for r_i, (k, v) in enumerate(sum_rows, start=1):
        ws_sum.cell(row=r_i, column=1, value=k)
        ws_sum.cell(row=r_i, column=2, value=v)

    if diag_rows:
        start_r = len(sum_rows) + 3
        ws_sum.cell(row=start_r, column=1, value="Diagnostics")
        headers = ["sheet","items","missing_rate_or_mapping","cols","size_row","mat_row","sides_row","qty_row","units","ds_loading_pct","total_sqm","subtotal"]
        for c_i, h in enumerate(headers, start=1):
            ws_sum.cell(row=start_r+1, column=c_i, value=h)
        for rr, d in enumerate(diag_rows, start=start_r+2):
            for c_i, h in enumerate(headers, start=1):
                ws_sum.cell(row=rr, column=c_i, value=d.get(h))

    ws_li = wb.create_sheet("Line Items")
    if len(all_df):
        cols = ["sheet","col_letter","qty","sides","ds_factor","shape","size_text","stock_customer","stock_std","sqm_each","total_sqm","sqm_rate","line_total"]
        cols = [c for c in cols if c in all_df.columns]
        for c_i, col in enumerate(cols, start=1):
            ws_li.cell(row=1, column=c_i, value=col)
        for r_i, row in enumerate(all_df[cols].itertuples(index=False), start=2):
            for c_i, v in enumerate(row, start=1):
                ws_li.cell(row=r_i, column=c_i, value=v)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


st.subheader("Save Quote to History (visible)")
with st.expander("Save to History", expanded=True):
    hh1, hh2, hh3, hh4 = st.columns([1.3, 1.3, 1.3, 3.1])
    hist_status = hh1.selectbox("Status", ["Pending","Won","Lost"], index=0, key="hist_status")
    hist_loss_reason = hh2.selectbox("Loss reason", ["", "Price", "Timing", "Spec/Capability", "Service", "Other"], index=0, key="hist_loss_reason")
    hist_sell_price = hh3.text_input("Sell price (optional)", value="", key="hist_sell_price")
    hist_notes = hh4.text_input("Notes", value="", key="hist_notes")

    if st.button("Save CURRENT quote/bundle to History", type="primary", key="hist_save_btn_top"):
        to_apply = st.session_state.bundle if st.session_state.bundle else {sheet_name: {"lines": lines.copy(), "settings": {}}}
        # compute totals from bundle
        all_rows = []
        for sh, data in to_apply.items():
            df_lines = data["lines"].copy()
            df_lines["sheet"] = sh
            all_rows.append(df_lines)
        all_df = pd.concat(all_rows, ignore_index=True) if all_rows else pd.DataFrame()
        total_sqm_all = float(pd.to_numeric(all_df.get("total_sqm", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if len(all_df) else 0.0
        subtotal_all = float(pd.to_numeric(all_df.get("line_total", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if len(all_df) else 0.0

        record = {
            "quote_id": str(uuid.uuid4()),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "customer": customer,
            "file_name": _uploaded_name,
            "status": hist_status,
            "loss_reason": hist_loss_reason,
            "sell_price": hist_sell_price,
            "notes": hist_notes,
            "price_row": int(st.session_state.get("global_price_row", price_row)),
            "ds_loading_pct": float(ds_loading_pct),
            "applied_sheets": list(to_apply.keys()),
            "total_sqm": total_sqm_all,
            "subtotal_material": subtotal_all,
            "bundle_settings": {sh: data.get("settings", {}) for sh, data in to_apply.items()},
            "bundle_json": serialize_bundle(to_apply),
        }
        save_history_record(record)
        st.success("Saved to history.")

b1, b2 = st.columns(2)
with b1:
    xbytes = export_preserving_excel_all_sheets()
    st.download_button(
        "Download Excel (ALL saved sheets)",
        data=xbytes,
        file_name=f"Quote_{customer}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
with b2:
    pbytes = export_quote_pdf(summary, review, title=f"Quote - {customer}")
    st.download_button(
        "Download Quote PDF",
        data=pbytes,
        file_name=f"Quote_{customer}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
        mime="application/pdf"
    )

with st.expander("Version / Debug", expanded=False):
    st.write("Last export apply mode:", st.session_state.get("apply_mode"))
    st.write("Selected sheets:", st.session_state.get("all_sheets_selected"))
    st.write("APP VERSION:", APP_VERSION)
    st.write("Columns processed:", f"{start_col_letter}:{end_col_letter}")
    st.write("Rows:", dict(size_row=int(size_row), material_row=int(mat_row), sides_row=int(sides_row), qty_row=int(qty_row), price_row=int(price_row)))
